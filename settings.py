import json
from pathlib import Path
import requests
import plotly.express as px
import openpyxl as oxl
import pandas as pd
import plotly.express as px
import streamlit as st
import plotly.graph_objects as go
from openpyxl.styles import Font, Color, Fill, PatternFill, Border, Side, numbers, Alignment
from openpyxl.utils import get_column_letter
import os
import base64

TITLE = "Databoard"
PAGE_ICON ="ico_potfolio.ico"


current_dir = Path(__file__).parent if "__file__" in locals() else Path.cwd()


# Credentials
config_Dash = current_dir / "hashed_pw.yaml"


# CSS
css_file = current_dir / "main.css"

pp_logo_portfolio = current_dir / "files" /  "logo_portfolio.png"
linkpic_code = current_dir / "files" /  "code.png"

# Tuto :
space = 15
tuto_space = 70
tuto_analyticdash_p = current_dir / "files" / "tuto_analyticdash.mp4"
tuto_analyticdash = str(tuto_analyticdash_p)

# lotties :
lottie_dashb = current_dir / "files" / "dashb.json"

def load_lottiefile(filepath : str):
    with open(filepath, "r") as f:
        return json.load(f)


def load_lottieurl(url: str):
    r = requests.get(url)
    if r.status_code != 200:
        return None
    return r.json()

# Clickable img
@st.cache_data
def get_base64_of_bin_file(bin_file):
    with open(bin_file, 'rb') as f:
        data = f.read()
    return base64.b64encode(data).decode()

@st.cache_data
def get_img_with_href(local_img_path, target_url, width, loc):
    img_format = os.path.splitext(local_img_path)[-1].replace('.', '')
    bin_str = get_base64_of_bin_file(local_img_path)
    html_code = f'''
        <a href="{target_url}" target="_{loc}" style="display: flex; justify-content: center; align-items: center;">
            <img src="data:image/{img_format};base64,{bin_str}" width="{width}" class="img-hover-effect">
        </a>'''
    return html_code

Excel_BdD = current_dir / "files" / "EXP_excel_data.xlsx"

@st.cache_data
def get_data_from_dataset(Excel_BdD):
    wb = oxl.load_workbook(Excel_BdD)
    wb_sheets = wb.sheetnames
    df = pd.read_excel(io=Excel_BdD, sheet_name=wb_sheets[0], header=0)
    return df


def total_kpi(df):
    tot_Achats = (df['Prix_Achat']*df['Qté_vendue']).sum()
    tot_ventes = df['Prix_vente_Total'].sum()
    tot_taxes = df['Taxe_5%'].sum()
    tot_profit_net= df['Profit_net'].sum()

    return [tot_Achats,tot_ventes,tot_taxes,tot_profit_net]


# graph volume des ventes / sexe 
def piechart_sales_gender(df):
    df_grouped_arr = df.groupby(by=['Gender']).sum()[['Prix_vente_Total']]

    pie_chart = px.pie(df_grouped_arr,
        values='Prix_vente_Total',
        names=df_grouped_arr.index,
        color=df_grouped_arr.index,
        color_discrete_map = {
            'Female': '#D67FD4',
            'Male': '#4572B3'})

    # Customize the pie chart
    pie_chart.update_traces(
        textposition='auto', #'inside', 'outside' 
        textinfo='percent+label')

    pie_chart.update_layout(
        font=dict(size=12, family='Arial Black'),
        legend=dict(
            font=dict(size=12, family='Arial Black'),
            x=0.1,
            y=-0.1,
            orientation='h'),
        width = 400,
        height = 400)

    return pie_chart


# graph volume des ventes / ville 
def piechart_sales_city(df):
    df_grouped_arr = df.groupby(by=['City']).sum()[['Prix_vente_Total']]

    pie_chart = px.pie(df_grouped_arr,
        values='Prix_vente_Total',
        names=df_grouped_arr.index,
        color=df_grouped_arr.index,
        color_discrete_map = {
            'Casablanca': '#D1E04F',
            'Rabat': '#53E04F',
            'Marrakech': '#9B3445'})

    # Customize the pie chart
    pie_chart.update_traces(
        textposition='auto', #'inside', 'outside' 
        textinfo='percent+label')

    pie_chart.update_layout(
        font=dict(size=12, family='Arial Black'),
        legend=dict(
            font=dict(size=12, family='Arial Black'),
            x=0.25,
            y=-0.1,
            orientation='h'),
        width = 400,
        height = 400)

    return pie_chart


# graph volume des ventes / pay_type 
@st.cache_data
def bar_chart_pay_type(df):
    
    df_grouped = df.groupby(by=['Payment_type']).sum()[['Prix_vente_Total']]
    color_map = {
        'Cash': 'green',
        'Credit card': '#94362D',
        'Phone wallet': '#2D5C94'}

    bar_chart = px.bar(df_grouped, 
        x=df_grouped.index, 
        y="Prix_vente_Total",
        color=df_grouped.index,
        color_discrete_map=color_map, )

    bar_chart.update_layout(
        xaxis_title=None,  # Set the x-axis title
        yaxis_title="Sales",  # Set the y-axis title
        font=dict(size=12, family='Arial Black'),
        legend=dict(title=None,x=0.1, y=1.1, orientation='h'),
        width = 450,
        height = 400,
        bargap=0.6)
    return bar_chart


# Evolution 
def get_data_sem_chart(df):

    df_sem = df.groupby(["Semaine"], as_index=False).agg({
    'Prix_vente_Total' : 'mean',
    'Profit_net': 'mean',
    'Qté_vendue':'mean'
    }).round(2)

    # Create the figure 
    trace1_0 = go.Scatter(x=df_sem['Semaine'], y=df_sem['Prix_vente_Total'], name="Total Sales",
        line=dict(width=2, color='#B4339D', dash='dash'), marker=dict(size=7, color='#B4339D'))

    trace1_1 = go.Scatter(x=df_sem['Semaine'], y=df_sem['Profit_net'], name="Net Profit",
        line=dict(width=2, color='#4088CD', dash='dash'), marker=dict(size=7, color='#4088CD'))

    trace2_1 = go.Scatter(x=df_sem['Semaine'], y=df_sem['Qté_vendue'], yaxis='y2',name="Quantity sold (Nbr)",
        line=dict(width=3, color='#7638C6'), marker=dict(size=10, color='#7638C6'))

    # Create a figure and add the traces to it
    fig = go.Figure()
    fig.add_traces([trace1_0, trace1_1, trace2_1])

    # Update the layout to include two y-axes
    fig.update_layout(yaxis=dict(title="Total Sales"),
        yaxis2=dict(title="Net Profit", overlaying='y', side='right'),
        xaxis=dict(tickvals=df_sem['Semaine'], ticktext=df_sem['Semaine']), # Affcher toutes les valeurs
        font=dict(size=12, family='Arial Black'),
        legend=dict(orientation="h", x=0.1, y=1.1, font=dict(size=12,family='Arial Black')))

    return fig


# Extract Data
def get_showndf(df):
    ladt_df = df.loc[:,['Semaine', 'City', "Gender", 'Type_produit', 'Prix_Vente', 'Qté_vendue',
        "Prix_vente_Total", 'Taxe_5%', "Profit_net", "Payment_type"]]

    ladt_df.sort_values(["Semaine", 'Prix_vente_Total'], ascending=[True,False], inplace=True)
    ladt_df.reset_index(drop=True, inplace=True)

    return ladt_df


# Curstom Excel file
def curstom_excel_df(excel_file):
    wb = oxl.load_workbook(excel_file)

    # Backgrounds
    Gris_fonce = PatternFill(patternType='solid', fgColor="0A0A0A") # Gris foncé
    Marron_Clair = PatternFill(patternType='solid', fgColor="C4BD97") # Marron clair

    # Borders
    side1 = Side(border_style='thin', color="000000")
    border1 = Border(top=side1 , bottom=side1, right=side1, left=side1)

    # Fontstyles
    font_style = Font(name= 'Calibri', size = 11, color= "000000", bold = True)

    # Alignements
    alignment_left = Alignment(horizontal="left",vertical="center")

    """-----------------------------Mise en forme Feuilles -----------------------------""" 
    ws = wb[wb.sheetnames[0]]
    max_c = ws.max_column # le max de cols (exploités) dans la feuille
    max_r = ws.max_row # le max de lignes (exploités) dans la feuille

    cell_fil1 = get_column_letter(max_c)
    ws.auto_filter.ref = f'A1:{cell_fil1}1' # Add a filter
    
    # Mise en forme cadre du tableau en gris :
    for row in ws.iter_rows(min_row=1, min_col=1, max_row=max_r+10, max_col = max_c+10):
        for cell in row:
            cell.fill=Gris_fonce

    # Mise en forme en-tete :
    for cell in ws[1]:
        cell.font=font_style

    # Hauteur de la 1ere ligne
    ws.row_dimensions[1].height = 30

    for row in ws.iter_rows(min_row=1, max_row=max_r, min_col=1, max_col = max_c):
        for cell in row:
            cell.fill=Marron_Clair
            cell.border=border1
            cell.alignment = alignment_left


    for index_col in range(1, max_c+1):
        i = get_column_letter(index_col)
        ws.column_dimensions[i].width = 13

    wb.save(excel_file)
    
    return excel_file
